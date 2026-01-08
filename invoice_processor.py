# invoice_processor.py
import pandas as pd
import os
from datetime import datetime
from pathlib import Path
import io
import warnings
from decimal import Decimal, ROUND_HALF_UP
warnings.filterwarnings('ignore')

class InvoiceProcessor:
    def __init__(self, brand_name, config):
        self.brand_name = brand_name
        self.config = config
        
    def process_invoices(self, uploaded_file):
        """Process uploaded file and return dictionary of invoices"""
        
        # Read uploaded file
        try:
            df = pd.read_excel(uploaded_file)
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")
        
        # Validate required columns
        required_cols = self.config['required_columns']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise Exception(f"Missing required columns: {missing_cols}")
        
        # Group by ASC (using ASC Name column)
        asc_column = self.config['asc_column']
        if asc_column not in df.columns:
            raise Exception(f"ASC column '{asc_column}' not found in data")
        
        asc_groups = df.groupby(asc_column)
        
        results = {}
        total_ascs = len(asc_groups)
        
        for i, (asc_name, asc_data) in enumerate(asc_groups, 1):
            # Generate invoice for this ASC
            invoice_bytes = self._generate_single_invoice(asc_name, asc_data)
            
            # Create supporting raw data file
            raw_data_bytes = self._create_supporting_data(asc_name, asc_data)
            
            # Calculate totals for summary
            total_amount = float(asc_data['Earning'].sum()) if 'Earning' in asc_data.columns else 0.0
            total_cod = float(asc_data['COD'].sum()) if 'COD' in asc_data.columns else 0.0
            
            results[asc_name] = {
                'invoice': invoice_bytes,
                'raw_data': raw_data_bytes,
                'records': len(asc_data),
                'total_amount': total_amount,
                'total_cod': total_cod,
                'invoice_number': f"INV-{datetime.now().strftime('%Y%m%d')}-{asc_name[:5]}"
            }
            
        return results
    
    def _generate_single_invoice(self, asc_name, asc_data):
        """Generate invoice for a single ASC"""
        
        # Extract ASC information (first record's details)
        first_record = asc_data.iloc[0]
        
        # For Candor CRM, don't include Owner Name and Contact No. in ASC details
        if self.brand_name == 'CandorCRM':
            asc_details_text = f"{asc_name}\n{first_record.get('Address', '')}"
        else:
            asc_details_text = f"{asc_name}\n{first_record.get('Address', '')}\nName: {first_record.get('Owner Name', '')} Mob. No.: {first_record.get('Contact No.', '')}"
        
        # Create invoice data structure
        invoice_data = {
            # ASC Details
            'asc_name': asc_name,
            'address': first_record.get('Address', ''),
            'owner_name': first_record.get('Owner Name', ''),
            'contact_no': first_record.get('Contact No.', ''),
            'pan_no': first_record.get('PAN No.', ''),
            'gst_no': first_record.get('GST No.', ''),
            
            # Invoice Details
            'invoice_number': first_record.get('Invoice Number', f"INV-{datetime.now().strftime('%Y%m%d')}-{asc_name[:5]}"),
            'invoice_date': datetime.now().strftime('%d-%b-%Y'),
            
            # Bill To (fixed for all brands)
            'bill_to': "RV Solutions Private Limited.\nD-59, Sector-2, Gautam Buddh Nagar, Noida,\nUttar Pradesh Noida-201301.",
            
            # Item details
            'items': self._extract_items(asc_data),
            
            # Month info (extract from data or use current)
            'invoice_month': self._extract_invoice_month(asc_data),
            
            # Brand-specific settings
            'brand': self.brand_name,
            
            # Check if it's a freelancer (for Harman and LifeLong)
            'is_freelancer': (self.brand_name == 'Harman' or self.brand_name == 'LifeLong') and 'Free Lancer' in str(asc_name),
            
            # Totals
            'totals': self._calculate_totals(asc_data, self.brand_name, asc_name)
        }
        
        # Generate Excel invoice
        return self._create_excel_invoice(asc_name, invoice_data)
    
    def _extract_items(self, asc_data):
        """Extract and group items based on brand"""
        items = []
        
        if self.brand_name == 'Amazon':
            # Amazon logic - group by category
            category_column = 'category'
            
            if category_column in asc_data.columns and not asc_data[category_column].isna().all():
                category_groups = asc_data.groupby(category_column)
                
                for category, group in category_groups:
                    category_str = str(category) if not pd.isna(category) else "Services"
                    
                    # Quantity calculation
                    if 'quantity' in group.columns:
                        total_qty = int(group['quantity'].sum())
                    else:
                        total_qty = len(group)
                    
                    # Amount calculation
                    total_earning = 0.0
                    if 'Earning' in group.columns:
                        total_earning = float(group['Earning'].sum())
                    
                    items.append({
                        'description': category_str,
                        'sac_code': '998715',
                        'quantity': total_qty,
                        'amount': round(total_earning, 2)
                    })
            else:
                # Fallback: single service row
                if 'quantity' in asc_data.columns:
                    total_qty = int(asc_data['quantity'].sum())
                else:
                    total_qty = len(asc_data)
                
                total_earning = 0.0
                if 'Earning' in asc_data.columns:
                    total_earning = float(asc_data['Earning'].sum())
                
                items.append({
                    'description': 'Services',
                    'sac_code': '998715',
                    'quantity': total_qty,
                    'amount': round(total_earning, 2)
                })
        
        elif self.brand_name == 'Harman':
            # Harman logic - group by Description column
            description_column = 'Description'
            
            if description_column in asc_data.columns and not asc_data[description_column].isna().all():
                description_groups = asc_data.groupby(description_column)
                
                for description, group in description_groups:
                    description_str = str(description) if not pd.isna(description) else "Services"
                    
                    # Quantity is the count of rows for this description
                    total_qty = len(group)
                    
                    # Amount from Call Charge column
                    total_amount = 0.0
                    if 'Call Charge' in group.columns:
                        total_amount = float(group['Call Charge'].sum())
                    
                    items.append({
                        'description': description_str,
                        'sac_code': '998715',
                        'quantity': total_qty,
                        'amount': round(total_amount, 2)
                    })
            else:
                # Fallback: single service row
                total_qty = len(asc_data)
                
                total_amount = 0.0
                if 'Call Charge' in asc_data.columns:
                    total_amount = float(asc_data['Call Charge'].sum())
                
                items.append({
                    'description': 'Services',
                    'sac_code': '998715',
                    'quantity': total_qty,
                    'amount': round(total_amount, 2)
                })
        
        elif self.brand_name == 'Philips':
            # Philips logic - group by Category column
            category_column = 'Category'
            
            if category_column in asc_data.columns and not asc_data[category_column].isna().all():
                category_groups = asc_data.groupby(category_column)
                
                for category, group in category_groups:
                    category_str = str(category) if not pd.isna(category) else "Services"
                    
                    # Quantity is the count of rows for this category
                    total_qty = len(group)
                    
                    # Amount from Final Amount column
                    total_amount = 0.0
                    if 'Final Amount' in group.columns:
                        total_amount = float(group['Final Amount'].sum())
                    
                    items.append({
                        'description': category_str,
                        'sac_code': '998715',
                        'quantity': total_qty,
                        'amount': round(total_amount, 2)
                    })
            else:
                # Fallback: single service row
                total_qty = len(asc_data)
                
                total_amount = 0.0
                if 'Final Amount' in asc_data.columns:
                    total_amount = float(asc_data['Final Amount'].sum())
                
                items.append({
                    'description': 'Services',
                    'sac_code': '998715',
                    'quantity': total_qty,
                    'amount': round(total_amount, 2)
                })
        
        elif self.brand_name == 'LifeLong':
            # LifeLong logic - group by Description column
            description_column = 'Description'
            
            if description_column in asc_data.columns and not asc_data[description_column].isna().all():
                description_groups = asc_data.groupby(description_column)
                
                for description, group in description_groups:
                    description_str = str(description) if not pd.isna(description) else "Services"
                    
                    # Quantity is the count of rows for this description
                    total_qty = len(group)
                    
                    # Amount from Final Amount column
                    total_amount = 0.0
                    if 'Final Amount' in group.columns:
                        total_amount = float(group['Final Amount'].sum())
                    
                    items.append({
                        'description': description_str,
                        'sac_code': '998715',
                        'quantity': total_qty,
                        'amount': round(total_amount, 2)
                    })
            else:
                # Fallback: single service row
                total_qty = len(asc_data)
                
                total_amount = 0.0
                if 'Final Amount' in asc_data.columns:
                    total_amount = float(asc_data['Final Amount'].sum())
                
                items.append({
                    'description': 'Services',
                    'sac_code': '998715',
                    'quantity': total_qty,
                    'amount': round(total_amount, 2)
                })
        
        elif self.brand_name == 'CandorCRM':
            # Candor CRM logic - group by Claim Status column
            claim_status_column = 'Claim Status'
            
            if claim_status_column in asc_data.columns and not asc_data[claim_status_column].isna().all():
                claim_groups = asc_data.groupby(claim_status_column)
                
                for claim_status, group in claim_groups:
                    claim_str = str(claim_status) if not pd.isna(claim_status) else "Services"
                    
                    # Quantity is the count of rows for this claim status
                    total_qty = len(group)
                    
                    # Calculate rate (average amount per item)
                    total_amount = 0.0
                    if 'Amount' in group.columns:
                        total_amount = float(group['Amount'].sum())
                    
                    # Calculate rate (average amount - for Candor CRM, rate is the individual item amount)
                    rate = 0.0
                    if total_qty > 0 and 'Amount' in group.columns:
                        # Get the first amount as rate (assuming same rate for all rows in same group)
                        rate = float(group['Amount'].iloc[0]) if len(group) > 0 else 0.0
                    
                    # For Candor CRM, we need to include sac_code in the dictionary
                    # even though it's not shown in the table headers
                    items.append({
                        'description': claim_str,
                        'sac_code': '998729',  # Fixed SAC code for Candor CRM
                        'quantity': total_qty,
                        'rate': round(rate, 2),
                        'amount': round(total_amount, 2)
                    })
            else:
                # Fallback: single service row
                total_qty = len(asc_data)
                
                total_amount = 0.0
                rate = 0.0
                if 'Amount' in asc_data.columns:
                    total_amount = float(asc_data['Amount'].sum())
                    if total_qty > 0:
                        rate = float(asc_data['Amount'].iloc[0]) if len(asc_data) > 0 else 0.0
                
                items.append({
                    'description': 'Services',
                    'sac_code': '998729',  # Fixed SAC code for Candor CRM
                    'quantity': total_qty,
                    'rate': round(rate, 2),
                    'amount': round(total_amount, 2)
                })
        
        return items
    
    def _extract_invoice_month(self, asc_data):
        """Extract month from data if available"""
        # Try to get from order_day or similar columns
        date_columns = ['order_day', 'invoice_date', 'appointment_start_time']
        for col in date_columns:
            if col in asc_data.columns:
                try:
                    # Get first date and extract month
                    sample_date = pd.to_datetime(asc_data[col].iloc[0])
                    return sample_date.strftime("%B %Y")
                except:
                    continue
        
        # Default to current month
        return datetime.now().strftime("%B %Y")
    
    def _calculate_totals(self, asc_data, brand_name=None, asc_name=""):
        """Calculate all totals with proper rounding"""
        # Use passed parameters or instance attributes
        if brand_name is None:
            brand_name = self.brand_name
            asc_name = asc_name if asc_name else ""
        
        def round_decimal(value):
            return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        # Check if it's a freelancer (for Harman and LifeLong)
        is_freelancer = (brand_name == 'Harman' or brand_name == 'LifeLong') and 'Free Lancer' in str(asc_name)
        
        if brand_name == 'Amazon':
            total_qty = asc_data['quantity'].sum() if 'quantity' in asc_data.columns else len(asc_data)
            total_amount = float(asc_data['Earning'].sum()) if 'Earning' in asc_data.columns else 0.0
            total_cod = float(asc_data['COD'].sum()) if 'COD' in asc_data.columns else 0.0
        elif brand_name == 'Harman':
            total_qty = len(asc_data)  # Total row count
            total_amount = float(asc_data['Call Charge'].sum()) if 'Call Charge' in asc_data.columns else 0.0
            total_cod = 0.0  # Harman doesn't have COD
        elif brand_name == 'Philips':
            total_qty = len(asc_data)  # Total row count
            total_amount = float(asc_data['Final Amount'].sum()) if 'Final Amount' in asc_data.columns else 0.0
            total_cod = 0.0  # Philips doesn't have COD
        elif brand_name == 'LifeLong':
            total_qty = len(asc_data)  # Total row count
            total_amount = float(asc_data['Final Amount'].sum()) if 'Final Amount' in asc_data.columns else 0.0
            total_cod = 0.0  # LifeLong doesn't have COD
        elif brand_name == 'Candor CRM':
            total_qty = len(asc_data)  # Total row count
            total_amount = float(asc_data['Amount'].sum()) if 'Amount' in asc_data.columns else 0.0
            total_cod = 0.0  # Candor CRM doesn't have COD
        else:
            total_qty = len(asc_data)
            total_amount = 0.0
            total_cod = 0.0
        
        # Convert to Decimal for precise calculations
        total_amount_dec = Decimal(str(total_amount))
        total_cod_dec = Decimal(str(total_cod))
        
        # Calculate GST with proper rounding
        if is_freelancer:
            igst_dec = Decimal('0.00')
        else:
            igst_dec = round_decimal(total_amount_dec * Decimal('0.18'))
        
        # Calculate invoice amount
        invoice_amount_dec = round_decimal(total_amount_dec + igst_dec)
        
        # For Harman, Philips, LifeLong and Candor CRM, net amount is same as invoice amount (no COD deduction)
        if brand_name in ['Harman', 'Philips', 'LifeLong', 'Candor CRM']:
            net_amount_dec = invoice_amount_dec
        else:
            net_amount_dec = round_decimal(invoice_amount_dec - total_cod_dec)
        
        # Convert back to float for the rest of the code
        totals = {
            'total_qty': int(total_qty),
            'total_amount': float(total_amount_dec),
            'total_cod': float(total_cod_dec),
            'igst': float(igst_dec),
            'cgst': 0.0,
            'sgst': 0.0,
            'invoice_amount': float(invoice_amount_dec),
            'net_amount': float(net_amount_dec),
            'is_freelancer': is_freelancer,
            'brand': brand_name
        }
        
        # Convert to words
        totals['amount_in_words'] = self._number_to_words(float(invoice_amount_dec))
        
        return totals
    
    def _number_to_words(self, num):
        """Convert number to Indian rupee words matching invoice format"""
        def convert_to_words(n):
            ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten",
                "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
            tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
            
            if n < 20:
                return ones[n]
            elif n < 100:
                return tens[n // 10] + (" " + ones[n % 10] if n % 10 != 0 else "")
            elif n < 1000:
                return ones[n // 100] + " Hundred" + (" and " + convert_to_words(n % 100) if n % 100 != 0 else "")
            elif n < 100000:
                return convert_to_words(n // 1000) + " Thousand" + (" " + convert_to_words(n % 1000) if n % 1000 != 0 else "")
            elif n < 10000000:
                return convert_to_words(n // 100000) + " Lakh" + (" " + convert_to_words(n % 100000) if n % 100000 != 0 else "")
            else:
                return convert_to_words(n // 10000000) + " Crore" + (" " + convert_to_words(n % 10000000) if n % 10000000 != 0 else "")
        
        # Ensure num is float
        num = float(num)
        
        # Separate rupees and paise
        rupees = int(num)
        paise = int(round((num - rupees) * 100))
        
        # Handle special case for exact rupees
        if paise == 0:
            return f"{convert_to_words(rupees)} Rupees Only"
        else:
            return f"{convert_to_words(rupees)} Rupees and {convert_to_words(paise)} Paise Only"
    
    def _create_excel_invoice(self, asc_name, invoice_data):
        """Create properly formatted Excel invoice in memory - returns bytes"""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
        
        # For Candor CRM, use a completely different format
        if invoice_data.get('brand') == 'CandorCRM':
            return self._create_candor_crm_invoice(asc_name, invoice_data)
        
        # Create a new workbook in memory
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"
        
        # Define styles
        bold_font = Font(bold=True)
        normal_font = Font()
        right_align = Alignment(horizontal='right', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        center_align = Alignment(horizontal='center', vertical='center')
        top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define no border style for spacing rows
        no_border = Border(
            left=Side(style='none'),
            right=Side(style='none'),
            top=Side(style='none'),
            bottom=Side(style='none')
        )
        
        # Define light peach color fill
        peach_fill = PatternFill(start_color='FFFFE5CC',
                                end_color='FFFFE5CC',
                                fill_type='solid')
        
        # ===== INVOICE TITLE BASED ON BRAND =====
        if invoice_data.get('brand') == 'Harman' or invoice_data.get('brand') == 'LifeLong':
            invoice_title = "Bill of Supply"
        elif invoice_data.get('brand') == 'Philips':
            invoice_title = "Tax Invoice"
        else:
            invoice_title = "Tax Invoice"

        ws.merge_cells('A1:D1')
        ws['A1'] = invoice_title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_align
        # Apply border and peach fill to all cells in merged range
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}1']
            cell.border = thin_border
            cell.fill = peach_fill
        
        # ===== ASC DETAILS SECTION =====
        asc_details_text = f"{invoice_data['asc_name']}\n{invoice_data['address']}\nName: {invoice_data['owner_name']} Mob. No.: {invoice_data['contact_no']}"
        
        ws.merge_cells('A2:B5')
        asc_details_cell = ws['A2']
        asc_details_cell.value = asc_details_text
        asc_details_cell.alignment = top_left_align
        # Apply border to all cells in merged range
        for row in range(2, 6):
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = thin_border
        
        # ===== INVOICE HEADER DETAILS (RIGHT SIDE) =====
        ws['C2'] = "Invoice Number:"
        ws['C2'].font = bold_font
        ws['D2'] = invoice_data['invoice_number']
        
        ws['C3'] = "Invoice Date:"
        ws['C3'].font = bold_font
        ws['D3'] = invoice_data['invoice_date']
        
        ws['C4'] = "PAN No.:"
        ws['C4'].font = bold_font
        ws['D4'] = invoice_data['pan_no']
        
        ws['C5'] = "GST No.:"
        ws['C5'].font = bold_font
        ws['D5'] = invoice_data['gst_no']
        
        # ===== BILL TO SECTION =====
        ws.merge_cells('A6:B9')
        bill_to_text = "Bill To,\nRV Solutions Private Limited.\nD-59, Sector-2, Gautam Buddh Nagar, Noida,\nUttar Pradesh Noida-201301."
        bill_to_cell = ws['A6']
        bill_to_cell.value = bill_to_text
        bill_to_cell.alignment = top_left_align
        # Apply border to all cells in merged range
        for row in range(6, 10):
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = thin_border
        
        # ===== COMPANY DETAILS (RIGHT SIDE OF BILL TO) =====
        ws['C6'] = "PAN No.:"
        ws['C6'].font = bold_font
        ws['D6'] = "AADCR9806P"
        
        ws['C7'] = "GST No.:"
        ws['C7'].font = bold_font
        ws['D7'] = "09AADCR9806PJZL"
        
        ws['C8'] = "State Code:"
        ws['C8'].font = bold_font
        ws['D8'] = "'09"
        
        ws['C9'] = "Place of Supply:"
        ws['C9'].font = bold_font
        ws['D9'] = "Uttar Pradesh"
        
        # Align all right side cells
        for row in range(2, 10):
            for col in [3, 4]:  # Columns C and D
                cell = ws.cell(row=row, column=col)
                cell.alignment = left_align
                if cell.value and ":" in str(cell.value):
                    cell.font = bold_font

        for row in range(2, 10):
            for col in ['C', 'D']:
                ws[f'{col}{row}'].border = thin_border
        
        # ===== MONTH TITLE - MERGED =====
        if invoice_data.get('brand') == 'Harman':
            month_title = f"Harman Invoice Month of {invoice_data['invoice_month']}"
        elif invoice_data.get('brand') == 'Philips':
            month_title = f"Philips Invoice Month of {invoice_data['invoice_month']}"
        elif invoice_data.get('brand') == 'LifeLong':
            month_title = f"LifeLong Invoice Month of {invoice_data['invoice_month']}"
        else:
            month_title = f"Amazon Invoice Month of {invoice_data['invoice_month']}"

        ws.merge_cells('A10:D10')
        ws['A10'] = month_title
        ws['A10'].font = Font(bold=True)
        ws['A10'].alignment = center_align
        # Apply border to all cells in merged range
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}10'].border = thin_border
        
        # ===== TABLE HEADERS WITH BORDERS AND PEACH FILL =====
        headers = ["Description", "SAC Code", "Qty", "Amount"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=11, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.fill = peach_fill  # Add peach fill to header row
            if header == "Amount":
                cell.alignment = right_align
            elif header == "Qty":
                cell.alignment = center_align
            else:
                cell.alignment = left_align
        
        # ===== ADD ITEMS =====
        current_row = 12
        items = invoice_data['items']
        
        for idx, item in enumerate(items):
            # Add item row
            desc_cell = ws.cell(row=current_row, column=1, value=item['description'])
            desc_cell.border = thin_border
            desc_cell.alignment = left_align
            
            sac_cell = ws.cell(row=current_row, column=2, value=item['sac_code'])
            sac_cell.border = thin_border
            sac_cell.alignment = center_align
            
            qty_cell = ws.cell(row=current_row, column=3, value=item['quantity'])
            qty_cell.border = thin_border
            qty_cell.alignment = right_align
            
            amount_cell = ws.cell(row=current_row, column=4, value=item['amount'])
            amount_cell.border = thin_border
            amount_cell.alignment = right_align
            amount_cell.number_format = '#,##0.00'
            
            current_row += 1
            
            # Add spacing rows after each item for LifeLong brand only
            if invoice_data.get('brand') == 'LifeLong' and idx < len(items) - 1:
                # Add 2-3 empty rows without borders for spacing
                spacing_rows = 3
                
                for _ in range(spacing_rows):
                    for col in range(1, 5):  # Columns A to D
                        cell = ws.cell(row=current_row, column=col)
                        cell.border = no_border  # No borders for spacing rows
                        cell.value = ""  # Empty value
                    current_row += 1
        
        # ===== TOTAL ROW =====
        total_row = current_row
        
        ws.merge_cells(f'A{total_row}:B{total_row}')
        total_label = ws.cell(row=total_row, column=1, value="")
        for col in ['A', 'B']:
            ws[f'{col}{total_row}'].border = thin_border
        
        total_qty_cell = ws.cell(row=total_row, column=3, value=invoice_data['totals']['total_qty'])
        total_qty_cell.border = thin_border
        total_qty_cell.font = bold_font
        total_qty_cell.alignment = right_align
        
        total_amount_cell = ws.cell(row=total_row, column=4, value=invoice_data['totals']['total_amount'])
        total_amount_cell.border = thin_border
        total_amount_cell.font = bold_font
        total_amount_cell.alignment = right_align
        total_amount_cell.number_format = '#,##0.00'
        
        # ===== GST AND TOTALS SECTION =====
        gst_start = total_row + 1
        
        # IGST Row
        ws.merge_cells(f'A{gst_start}:B{gst_start}')
        igst_label = ws.cell(row=gst_start, column=1, value="IGST")
        igst_label.font = bold_font
        igst_label.alignment = right_align
        for col in ['A', 'B']:
            ws[f'{col}{gst_start}'].border = thin_border
        
        if invoice_data['totals'].get('is_freelancer', False):
            ws.cell(row=gst_start, column=3, value="-").alignment = right_align
            ws.cell(row=gst_start, column=3).border = thin_border
            igst_value = "-"
        else:
            ws.cell(row=gst_start, column=3, value="18%").alignment = right_align
            ws.cell(row=gst_start, column=3).border = thin_border
            igst_value = invoice_data['totals']['igst']
        
        igst_cell = ws.cell(row=gst_start, column=4, value=igst_value)
        igst_cell.alignment = right_align
        if not invoice_data['totals'].get('is_freelancer', False):
            igst_cell.number_format = '#,##0.00'
        igst_cell.border = thin_border
        
        # CGST Row (show for both brands but with "-")
        ws.merge_cells(f'A{gst_start+1}:B{gst_start+1}')
        cgst_label = ws.cell(row=gst_start+1, column=1, value="CGST")
        cgst_label.font = bold_font
        cgst_label.alignment = right_align
        for col in ['A', 'B']:
            ws[f'{col}{gst_start+1}'].border = thin_border
        
        ws.cell(row=gst_start+1, column=3, value="-").alignment = center_align
        ws.cell(row=gst_start+1, column=3).border = thin_border
        ws.cell(row=gst_start+1, column=4, value="-").alignment = center_align
        ws.cell(row=gst_start+1, column=4).border = thin_border
        
        # SGST Row (show for both brands but with "-")
        ws.merge_cells(f'A{gst_start+2}:B{gst_start+2}')
        sgst_label = ws.cell(row=gst_start+2, column=1, value="SGST")
        sgst_label.font = bold_font
        sgst_label.alignment = right_align
        for col in ['A', 'B']:
            ws[f'{col}{gst_start+2}'].border = thin_border
        
        ws.cell(row=gst_start+2, column=3, value="-").alignment = center_align
        ws.cell(row=gst_start+2, column=3).border = thin_border
        ws.cell(row=gst_start+2, column=4, value="-").alignment = center_align
        ws.cell(row=gst_start+2, column=4).border = thin_border
        
        # Invoice Amount Row - WITH PEACH FILL
        invoice_row = gst_start + 3
        ws.merge_cells(f'A{invoice_row}:B{invoice_row}')
        invoice_label = ws.cell(row=invoice_row, column=1, value="Invoice Amount")
        invoice_label.font = bold_font
        invoice_label.alignment = right_align
        invoice_label.fill = peach_fill  # Add peach fill
        for col in ['A', 'B']:
            cell = ws[f'{col}{invoice_row}']
            cell.border = thin_border
            cell.fill = peach_fill
        
        ws.cell(row=invoice_row, column=3).border = thin_border
        ws.cell(row=invoice_row, column=3).fill = peach_fill  # Add peach fill
        invoice_amount_cell = ws.cell(row=invoice_row, column=4, value=invoice_data['totals']['invoice_amount'])
        invoice_amount_cell.font = bold_font
        invoice_amount_cell.alignment = right_align
        invoice_amount_cell.number_format = '#,##0.00'
        invoice_amount_cell.border = thin_border
        invoice_amount_cell.fill = peach_fill  # Add peach fill
        
        # Only show Advance Received (COD) and Net Amount for Amazon
        if invoice_data.get('brand') == 'Amazon':
            # Advance Received (COD) Row
            cod_row = gst_start + 4
            ws.merge_cells(f'A{cod_row}:B{cod_row}')
            cod_label = ws.cell(row=cod_row, column=1, value="Advance Received (COD)")
            cod_label.font = bold_font
            cod_label.alignment = right_align
            for col in ['A', 'B']:
                ws[f'{col}{cod_row}'].border = thin_border
            
            ws.cell(row=cod_row, column=3).border = thin_border
            cod_cell = ws.cell(row=cod_row, column=4, value=invoice_data['totals']['total_cod'])
            cod_cell.alignment = right_align
            cod_cell.number_format = '#,##0.00'
            cod_cell.border = thin_border
            
            # Net Amount Row
            net_row = gst_start + 5
            ws.merge_cells(f'A{net_row}:B{net_row}')
            net_label = ws.cell(row=net_row, column=1, value="Net Amount")
            net_label.font = bold_font
            net_label.alignment = right_align
            for col in ['A', 'B']:
                ws[f'{col}{net_row}'].border = thin_border
            
            ws.cell(row=net_row, column=3).border = thin_border
            net_cell = ws.cell(row=net_row, column=4, value=invoice_data['totals']['net_amount'])
            net_cell.font = bold_font
            net_cell.alignment = right_align
            net_cell.number_format = '#,##0.00'
            net_cell.border = thin_border
            
            words_start_row = net_row + 1
        else:
            # For Harman, Philips and LifeLong, skip COD and Net Amount rows, start words after Invoice Amount
            words_start_row = invoice_row + 1
        
        # ===== AMOUNT IN WORDS =====
        words_row = words_start_row
        
        ws.merge_cells(f'A{words_row}:D{words_row}')
        words_label = ws.cell(row=words_row, column=1, value="Invoice Amount (in words)")
        words_label.font = bold_font
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{words_row}'].border = thin_border
        
        ws.merge_cells(f'A{words_row+1}:D{words_row+1}')
        amount_words = ws.cell(row=words_row+1, column=1, value=invoice_data['totals']['amount_in_words'])
        amount_words.alignment = left_align
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{words_row+1}'].border = thin_border
        
        # ===== DECLARATION AND TERMS =====
        declaration_row = words_row + 2
        
        # Merge A:B for declaration
        ws.merge_cells(f'A{declaration_row}:B{declaration_row+8}')
        
        # Use brand-specific declaration
        if invoice_data.get('brand') == 'Harman' or invoice_data.get('brand') == 'LifeLong':
            declaration_text = "Declaration:- We declare that this invoice shows the actual price of the goods/services described and that all particulars are true and correct.\n\nTerms: * In case of non reflection of the GST amount in GSTR-2B of RV Solutions Pvt. Ltd. within 30th-June of Next Financial year, we agree to pay RV Solutions Pvt. Ltd. the GST amount along with interest @24% p.a. on delayed payment."
        else:
            # Amazon and Philips use the same declaration
            declaration_text = "Declaration:- We declare that this invoice shows the actual price of the goods/services described and that all particulars are true and correct.\n\n* In case of non reflection of the GST amount in GSTR-2B of RV Solutions Pvt. Ltd. within 30th-June of Next Financial year, we agree to pay RV Solutions Pvt. Ltd. the GST amount along with interest @18% p.a. on delayed payment."
        
        declaration_cell = ws.cell(row=declaration_row, column=1, value=declaration_text)
        declaration_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        for row in range(declaration_row, declaration_row + 9):
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = thin_border
        
        # ===== AUTHORIZED SIGNATORY =====
        sign_start_row = declaration_row
        sign_end_row = declaration_row + 8
        
        ws.merge_cells(f'C{sign_start_row}:D{sign_end_row}')
        
        signatory_cell = ws.cell(
            row=sign_start_row,
            column=3,
            value="Authorised Signatory"
        )
        signatory_cell.font = bold_font
        signatory_cell.alignment = Alignment(
            horizontal='center',
            vertical='bottom'
        )
        
        for row in range(sign_start_row, sign_end_row + 1):
            for col in ['C', 'D']:
                ws[f'{col}{row}'].border = thin_border
        
        # ===== ADJUST COLUMN WIDTHS =====
        column_widths = {
            'A': 35,   # Description
            'B': 10,   # SAC Code
            'C': 16,   # Qty/Percentage
            'D': 18,   # Amount
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # ===== SAVE TO BYTES =====
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_candor_crm_invoice(self, asc_name, invoice_data):
        """Create Candor CRM specific invoice format"""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
        
        # Create a new workbook in memory
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"
        
        # Define styles
        bold_font = Font(bold=True)
        normal_font = Font()
        right_align = Alignment(horizontal='right', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        center_align = Alignment(horizontal='center', vertical='center')
        top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define light peach color fill
        peach_fill = PatternFill(start_color='FFFFE5CC',
                                end_color='FFFFE5CC',
                                fill_type='solid')
        
        # ===== TAX INVOICE TITLE =====
        ws.merge_cells('A1:D1')
        ws['A1'] = "Tax Invoice"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_align
        # Apply border and peach fill to all cells in merged range
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}1']
            cell.border = thin_border
            cell.fill = peach_fill
        
        # ===== ASC DETAILS SECTION (No Owner Name or Mob No for Candor CRM) =====
        asc_details_text = f"{invoice_data['asc_name']}\n{invoice_data['address']}"
        
        ws.merge_cells('A2:B5')
        asc_details_cell = ws['A2']
        asc_details_cell.value = asc_details_text
        asc_details_cell.alignment = top_left_align
        # Apply border to all cells in merged range
        for row in range(2, 6):
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = thin_border
        
        # ===== INVOICE HEADER DETAILS (RIGHT SIDE) =====
        ws['C2'] = "Invoice Number:"
        ws['C2'].font = bold_font
        ws['D2'] = invoice_data['invoice_number']
        
        ws['C3'] = "Invoice Date:"
        ws['C3'].font = bold_font
        ws['D3'] = invoice_data['invoice_date']
        
        ws['C4'] = "PAN No.:"
        ws['C4'].font = bold_font
        ws['D4'] = invoice_data['pan_no']
        
        ws['C5'] = "GST No.:"
        ws['C5'].font = bold_font
        ws['D5'] = invoice_data['gst_no']
        
        # ===== SAC CODE ROW (Additional row for Candor CRM) =====
        ws['C6'] = "SAC Code:"
        ws['C6'].font = bold_font
        ws['D6'] = "998729"
        
        # ===== BUYER SECTION (Instead of Bill To) =====
        ws.merge_cells('A7:B10')
        buyer_text = "Buyer\nRV Solutions Pvt. Ltd.\nD-59, Sector-2, District-Gautam Buddh Nagar, Noida,\nUttar Pradesh - 201301.\nContact No.-8588881737"
        buyer_cell = ws['A7']
        buyer_cell.value = buyer_text
        buyer_cell.alignment = top_left_align
        buyer_cell.font = bold_font  # "Buyer" in bold
        # Apply border to all cells in merged range
        for row in range(7, 11):
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = thin_border
        
        # ===== COMPANY DETAILS (RIGHT SIDE OF BUYER) =====
        ws['C7'] = "PAN No.:"
        ws['C7'].font = bold_font
        ws['D7'] = "AADCR9806P"
        
        ws['C8'] = "GST No.:"
        ws['C8'].font = bold_font
        ws['D8'] = "09AADCR9806PJZL"
        
        ws['C9'] = "State Code:"
        ws['C9'].font = bold_font
        ws['D9'] = "'09"
        
        ws['C10'] = "Place of Supply:"
        ws['C10'].font = bold_font
        ws['D10'] = "Uttar Pradesh"
        
        # Align all right side cells
        for row in range(2, 11):
            for col in [3, 4]:  # Columns C and D
                cell = ws.cell(row=row, column=col)
                cell.alignment = left_align
                if cell.value and ":" in str(cell.value):
                    cell.font = bold_font

        for row in range(2, 11):
            for col in ['C', 'D']:
                ws[f'{col}{row}'].border = thin_border
        
        # ===== MONTH TITLE - MERGED (Candor CRM specific) =====
        month_title = f"Honor/Acwo Claim Month of {invoice_data['invoice_month']}"
        
        ws.merge_cells('A11:D11')
        ws['A11'] = month_title
        ws['A11'].font = Font(bold=True)
        ws['A11'].alignment = center_align
        # Apply border to all cells in merged range
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}11'].border = thin_border
        
        # ===== TABLE HEADERS WITH BORDERS AND PEACH FILL =====
        # Different headers for Candor CRM
        headers = ["Description", "Quantity", "Rate", "Amount"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=12, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.fill = peach_fill  # Add peach fill to header row
            if header == "Amount" or header == "Rate":
                cell.alignment = right_align
            elif header == "Quantity":
                cell.alignment = center_align
            else:
                cell.alignment = left_align
        
        # ===== ADD ITEMS (Different format for Candor CRM) =====
        current_row = 13
        for item in invoice_data['items']:
            # Description (Column A)
            desc_cell = ws.cell(row=current_row, column=1, value=item['description'])
            desc_cell.border = thin_border
            desc_cell.alignment = left_align
            
            # Quantity (Column B)
            qty_cell = ws.cell(row=current_row, column=2, value=item['quantity'])
            qty_cell.border = thin_border
            qty_cell.alignment = right_align
            
            # Rate (Column C)
            rate_cell = ws.cell(row=current_row, column=3, value=item['rate'])
            rate_cell.border = thin_border
            rate_cell.alignment = right_align
            rate_cell.number_format = '#,##0.00'
            
            # Amount (Column D) = Quantity * Rate
            amount_cell = ws.cell(row=current_row, column=4, value=item['amount'])
            amount_cell.border = thin_border
            amount_cell.alignment = right_align
            amount_cell.number_format = '#,##0.00'
            
            current_row += 1
        
        # ===== TOTAL ROW =====
        total_row = current_row
        
        ws.merge_cells(f'A{total_row}:C{total_row}')
        total_label = ws.cell(row=total_row, column=1, value="Total")
        total_label.font = bold_font
        total_label.alignment = right_align
        for col in ['A', 'B', 'C']:
            ws[f'{col}{total_row}'].border = thin_border
        
        total_amount_cell = ws.cell(row=total_row, column=4, value=invoice_data['totals']['total_amount'])
        total_amount_cell.border = thin_border
        total_amount_cell.font = bold_font
        total_amount_cell.alignment = right_align
        total_amount_cell.number_format = '#,##0.00'
        
        # ===== GST AND TOTALS SECTION =====
        gst_start = total_row + 1
        
        # IGST Row (18% for Candor CRM)
        ws.merge_cells(f'A{gst_start}:C{gst_start}')
        igst_label = ws.cell(row=gst_start, column=1, value="IGST")
        igst_label.font = bold_font
        igst_label.alignment = right_align
        for col in ['A', 'B', 'C']:
            ws[f'{col}{gst_start}'].border = thin_border
        
        ws.cell(row=gst_start, column=4, value="18%").alignment = right_align
        ws.cell(row=gst_start, column=4).border = thin_border
        
        # IGST Amount Row
        ws.merge_cells(f'A{gst_start+1}:C{gst_start+1}')
        igst_amount_label = ws.cell(row=gst_start+1, column=1, value="")
        igst_amount_label.alignment = right_align
        for col in ['A', 'B', 'C']:
            ws[f'{col}{gst_start+1}'].border = thin_border
        
        igst_amount_cell = ws.cell(row=gst_start+1, column=4, value=invoice_data['totals']['igst'])
        igst_amount_cell.alignment = right_align
        igst_amount_cell.number_format = '#,##0.00'
        igst_amount_cell.border = thin_border
        
        # CGST Row (show as "-")
        ws.merge_cells(f'A{gst_start+2}:C{gst_start+2}')
        cgst_label = ws.cell(row=gst_start+2, column=1, value="CGST")
        cgst_label.font = bold_font
        cgst_label.alignment = right_align
        for col in ['A', 'B', 'C']:
            ws[f'{col}{gst_start+2}'].border = thin_border
        
        ws.cell(row=gst_start+2, column=4, value="-").alignment = center_align
        ws.cell(row=gst_start+2, column=4).border = thin_border
        
        # SGST Row (show as "-")
        ws.merge_cells(f'A{gst_start+3}:C{gst_start+3}')
        sgst_label = ws.cell(row=gst_start+3, column=1, value="SGST")
        sgst_label.font = bold_font
        sgst_label.alignment = right_align
        for col in ['A', 'B', 'C']:
            ws[f'{col}{gst_start+3}'].border = thin_border
        
        ws.cell(row=gst_start+3, column=4, value="-").alignment = center_align
        ws.cell(row=gst_start+3, column=4).border = thin_border
        
        # GRAND TOTAL Row (Instead of "Invoice Amount") - WITH PEACH FILL
        grand_total_row = gst_start + 4
        ws.merge_cells(f'A{grand_total_row}:C{grand_total_row}')
        grand_total_label = ws.cell(row=grand_total_row, column=1, value="Grand Total")
        grand_total_label.font = bold_font
        grand_total_label.alignment = right_align
        grand_total_label.fill = peach_fill  # Add peach fill
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{grand_total_row}']
            cell.border = thin_border
            cell.fill = peach_fill
        
        grand_total_cell = ws.cell(row=grand_total_row, column=4, value=invoice_data['totals']['invoice_amount'])
        grand_total_cell.font = bold_font
        grand_total_cell.alignment = right_align
        grand_total_cell.number_format = '#,##0.00'
        grand_total_cell.border = thin_border
        grand_total_cell.fill = peach_fill  # Add peach fill
        
        # ===== AMOUNT IN WORDS =====
        words_row = grand_total_row + 1
        
        ws.merge_cells(f'A{words_row}:D{words_row}')
        words_label = ws.cell(row=words_row, column=1, value="Invoice Amount (in words)")
        words_label.font = bold_font
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{words_row}'].border = thin_border
        
        ws.merge_cells(f'A{words_row+1}:D{words_row+1}')
        amount_words = ws.cell(row=words_row+1, column=1, value=invoice_data['totals']['amount_in_words'])
        amount_words.alignment = left_align
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{words_row+1}'].border = thin_border
        
        # ===== DECLARATION AND TERMS =====
        declaration_row = words_row + 2
        
        # Merge A:B for declaration
        ws.merge_cells(f'A{declaration_row}:B{declaration_row+8}')
        
        # Use Amazon/Philips declaration for Candor CRM (18% interest)
        declaration_text = "Declaration:- We declare that this invoice shows the actual price of the goods/services described and that all particulars are true and correct.\n\n* In case of non reflection of the GST amount in GSTR-2B of RV Solutions Pvt. Ltd. within 30th-June of Next Financial year, we agree to pay RV Solutions Pvt. Ltd. the GST amount along with interest @18% p.a. on delayed payment."
        
        declaration_cell = ws.cell(row=declaration_row, column=1, value=declaration_text)
        declaration_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        for row in range(declaration_row, declaration_row + 9):
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = thin_border
        
        # ===== AUTHORIZED SIGNATORY =====
        sign_start_row = declaration_row
        sign_end_row = declaration_row + 8
        
        ws.merge_cells(f'C{sign_start_row}:D{sign_end_row}')
        
        signatory_cell = ws.cell(
            row=sign_start_row,
            column=3,
            value="Authorised Signatory"
        )
        signatory_cell.font = bold_font
        signatory_cell.alignment = Alignment(
            horizontal='center',
            vertical='bottom'
        )
        
        for row in range(sign_start_row, sign_end_row + 1):
            for col in ['C', 'D']:
                ws[f'{col}{row}'].border = thin_border
        
        # ===== ADJUST COLUMN WIDTHS =====
        column_widths = {
            'A': 35,   # Description
            'B': 12,   # Quantity
            'C': 12,   # Rate
            'D': 18,   # Amount
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # ===== SAVE TO BYTES =====
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _create_supporting_data(self, asc_name, asc_data):
        """Create supporting raw data file - returns bytes"""
        output = io.BytesIO()
        asc_data.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)
        return output.getvalue()